"""
ESPACE VIRTUELLE MINIÈRE DE SMC
Module de parsing des fichiers de log géologique (RC, Aircore, Diamond Drilling, Auger, Structural)

Hypothèse de structure (gabarits SMC) :
  - Feuille "Geologie" : ligne 1 = en-têtes de groupe, ligne 2 = en-têtes de colonnes,
    lignes 3+ = données. Colonnes :
    EASTING(manuel), NORTHING, ELEVATION, EASTING(repiquetage), NORTHING, ELEVATION,
    Sondage, From, To, Priorité, Lithologie, Code_Litho, Couleur, Grain, Texture,
    Oxydation, Magn, Durete, Contact, Code_Strat, Formation, Age, Type_Ech
  - Feuilles annexes W (Weathering), M (Mineralisation), V (Veines), Al (Alteration),
    Ox (Oxydation), Water : alignées ligne à ligne avec la feuille "Geologie"
    (même ordre d'intervalles From-To).
"""
import pandas as pd
import numpy as np
import openpyxl
import io
import hashlib

# ---------------------------------------------------------------------------
# Palette de couleurs lithologiques (HTML) - étendue, pas seulement du bleu
# ---------------------------------------------------------------------------
LITHO_PALETTE = [
    "#E07B39", "#3B7A57", "#C9A227", "#8E44AD", "#2980B9", "#D35400",
    "#27AE60", "#C0392B", "#16A085", "#7F8C8D", "#F39C12", "#2C3E50",
    "#E74C3C", "#9B59B6", "#1ABC9C", "#D4AC0D", "#A04000", "#229954",
    "#5DADE2", "#CB4335", "#76448A", "#196F3D", "#B9770E", "#34495E",
    "#F1948A", "#85C1E9", "#82E0AA", "#F8C471", "#BB8FCE", "#73C6B6",
]

WEATHERING_ORDER = ["SUP", "LAT", "SAP", "SAPROLITE", "SAPROCK", "FRESH", "BED", "BEDROCK"]

WEATHERING_COLORS = {
    "SUP": "#A9744F",       # Sol / superficiel
    "LAT": "#B5421E",       # Latérite
    "SAP": "#D2A679",       # Saprolite
    "SAPROLITE": "#D2A679",
    "SAPROCK": "#8C7853",   # Saprock
    "FRESH": "#5A5A5A",     # Roche fraîche / bedrock
    "BED": "#5A5A5A",
    "BEDROCK": "#5A5A5A",
}


def _stable_color(label, used_map):
    """Assigne une couleur stable et distincte à un libellé lithologique."""
    if label is None:
        label = "Indéterminé"
    label = str(label).strip()
    if label in used_map:
        return used_map[label]
    idx = len(used_map) % len(LITHO_PALETTE)
    color = LITHO_PALETTE[idx]
    used_map[label] = color
    return color


def build_litho_color_map(labels):
    used = {}
    for l in labels:
        _stable_color(l, used)
    return used


def classify_weathering(code_strat, formation):
    """Détermine la classe d'altération météorique (Latérite/Saprolite/Saprock/Bedrock)
    à partir des champs Code_Strat / Formation."""
    text = f"{code_strat or ''} {formation or ''}".upper()
    if "LAT" in text:
        return "LAT"
    if "SAPROC" in text:
        return "SAPROCK"
    if "SAP" in text:
        return "SAP"
    if "SUP" in text or "SOL" in text:
        return "SUP"
    if "FRA" in text or "FRESH" in text or "SAIN" in text or "BED" in text:
        return "FRESH"
    return "INDETERMINE"


def _read_sheet_raw(file_bytes, sheet_name):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    return rows


def parse_geologie_workbook(file_bytes, drill_type="RC"):
    """Parse un classeur de log (RC/AC/DD) avec feuilles Geologie, W, M, V, Al, Ox, Water.
    Retourne un DataFrame fusionné, une ligne par intervalle, avec toutes les infos
    géologiques, structurales, de minéralisation et d'altération."""
    rows = _read_sheet_raw(file_bytes, "Geologie")
    if rows is None or len(rows) < 3:
        return pd.DataFrame()

    header = rows[1]
    data_rows = rows[2:]
    cols = [
        "Easting_man", "Northing_man", "Elevation_man",
        "Easting_rep", "Northing_rep", "Elevation_rep",
        "Sondage", "From", "To", "Priorite", "Lithologie", "Code_Litho",
        "Couleur", "Grain", "Texture", "Oxydation", "Magn", "Durete",
        "Contact", "Code_Strat", "Formation", "Age", "Type_Ech",
    ]
    n = min(len(cols), len(header) if header else len(cols))
    df = pd.DataFrame(data_rows, columns=[cols[i] if i < len(cols) else f"col{i}" for i in range(len(data_rows[0]) if data_rows else len(cols))])
    df = df.dropna(how="all")
    if df.empty:
        return df

    # forward-fill du sondage + coordonnées collar (souvent renseignées uniquement à la 1ère ligne du trou)
    df["Sondage"] = df["Sondage"].ffill()
    for c in ["Easting_man", "Northing_man", "Elevation_man", "Easting_rep", "Northing_rep", "Elevation_rep"]:
        if c in df.columns:
            df[c] = df.groupby("Sondage")[c].transform(lambda s: s.ffill().bfill())

    df["Easting"] = df["Easting_rep"].combine_first(df["Easting_man"]) if "Easting_rep" in df.columns else df.get("Easting_man")
    df["Northing"] = df["Northing_rep"].combine_first(df["Northing_man"]) if "Northing_rep" in df.columns else df.get("Northing_man")
    df["Elevation"] = df["Elevation_rep"].combine_first(df["Elevation_man"]) if "Elevation_rep" in df.columns else df.get("Elevation_man")

    df["From"] = pd.to_numeric(df["From"], errors="coerce")
    df["To"] = pd.to_numeric(df["To"], errors="coerce")

    # fusion des feuilles annexes (alignement par position de ligne)
    for sheet, prefix in [("W", "W_"), ("M", "M_"), ("V", "V_"), ("Al", "Al_"), ("Ox", "Ox_"), ("Water", "Water_")]:
        raw = _read_sheet_raw(file_bytes, sheet)
        if raw is None or len(raw) < 2:
            continue
        hdr_row_idx = 1 if len(raw) > 1 else 0
        sub_header = raw[hdr_row_idx]
        sub_data = raw[hdr_row_idx + 1:]
        if not sub_data:
            continue
        ncols = max(len(r) for r in sub_data)
        sub_df = pd.DataFrame(sub_data, columns=[f"{prefix}c{i}" for i in range(ncols)])
        sub_df = sub_df.reindex(range(len(df))).reset_index(drop=True)
        df = df.reset_index(drop=True)
        df = pd.concat([df, sub_df], axis=1)

    df["Weathering_Class"] = df.apply(lambda r: classify_weathering(r.get("Code_Strat"), r.get("Formation")), axis=1)
    df["Drill_Type"] = drill_type
    df["Source_File_Hash"] = hashlib.md5(file_bytes).hexdigest()[:8]

    # indicateurs minéralisation / alteration (présence de % sulfure ou code alteration rempli)
    def has_value(row, cols_):
        for c in cols_:
            v = row.get(c)
            if v not in (None, "", " "):
                return True
        return False

    m_cols = [c for c in df.columns if c.startswith("M_")]
    al_cols = [c for c in df.columns if c.startswith("Al_")]
    v_cols = [c for c in df.columns if c.startswith("V_")]
    df["Has_Mineralisation"] = df.apply(lambda r: has_value(r, m_cols), axis=1) if m_cols else False
    df["Has_Alteration"] = df.apply(lambda r: has_value(r, al_cols), axis=1) if al_cols else False
    df["Has_Veine"] = df.apply(lambda r: has_value(r, v_cols), axis=1) if v_cols else False

    return df


def parse_auger_workbook(file_bytes):
    """Parse Log_Geochimie_Sols_AG.xlsx (feuille Logs_Sols) - une ligne = un intervalle, déjà
    avec Sondage par ligne. Contient les teneurs Au (Labo_Au_ppb)."""
    rows = _read_sheet_raw(file_bytes, "Logs_Sols")
    if rows is None or len(rows) < 2:
        return pd.DataFrame()
    header = rows[1]
    data_rows = rows[2:]
    cols = [
        "Easting", "Northing", "Elevation", "Sondage", "From", "To", "Interval_m",
        "Litho_Code", "Lithologie", "Code_Strat", "Formation", "Horizon_Sol", "Couleur",
        "Texture", "Structure_Sol", "Alteration", "Oxydation", "Ferruginisation",
        "Nodules", "Graviers_pct", "Humidite", "Consistance", "Commentaire",
        "Echant_ID", "Labo_Au_ppb", "Labo_Multi_ppm",
    ]
    if not data_rows:
        return pd.DataFrame()
    width = max(len(r) for r in data_rows)
    use_cols = (cols + [f"col{i}" for i in range(len(cols), width)])[:width]
    df = pd.DataFrame(data_rows, columns=use_cols)
    df = df.dropna(how="all")
    df["Sondage"] = df["Sondage"].ffill()
    for c in ["Easting", "Northing", "Elevation"]:
        if c in df.columns:
            df[c] = df.groupby("Sondage")[c].transform(lambda s: s.ffill().bfill())
    df["From"] = pd.to_numeric(df.get("From_m", df.get("From")), errors="coerce") if "From_m" not in df.columns else pd.to_numeric(df["From_m"], errors="coerce")
    if "From" not in df.columns:
        df["From"] = pd.to_numeric(df.get("From"), errors="coerce")
    df["To"] = pd.to_numeric(df.get("To_m", df.get("To")), errors="coerce") if "To_m" not in df.columns else pd.to_numeric(df["To_m"], errors="coerce")
    df["Weathering_Class"] = df.apply(lambda r: classify_weathering(r.get("Code_Strat"), r.get("Formation")), axis=1)
    df["Au_ppb"] = pd.to_numeric(df.get("Labo_Au_ppb"), errors="coerce")
    df["Drill_Type"] = "AUGER"
    return df


def parse_structural_workbook(file_bytes):
    """Parse logging_structural.xlsx (feuille Logs_Geotechniques)."""
    rows = _read_sheet_raw(file_bytes, "Logs_Geotechniques")
    if rows is None or len(rows) < 2:
        return pd.DataFrame()
    header = rows[1]
    data_rows = rows[2:]
    cols = [
        "Easting_man", "Northing_man", "Elevation_man", "Easting_rep", "Northing_rep", "Elevation_rep",
        "Sondage", "From_m", "To_m", "Epaisseur_m", "Litho", "Veine_Struct", "Foliation_structure",
        "Fracture_structure", "Bande_alteration_structure", "Stringer_structure", "Shear_veine_structure",
        "Alpha_deg", "Beta_deg", "Fiabilite_orientation", "Qualite", "OCTypes", "Azimut", "Pendage",
        "RQD_auto", "GSI_auto", "Intensite", "Frequence", "Espacement_m", "Ouverture_mm", "Remplissage",
        "Py_sulfure", "As_sulfure", "Cu_sulfure", "Cha_sulfure", "Pourcentage_sulfures_total",
        "Top_ou_Bottom", "Eau", "Commentaire_Structurales", "Commentaires_GSI",
    ]
    if not data_rows:
        return pd.DataFrame()
    width = max(len(r) for r in data_rows)
    use_cols = (cols + [f"col{i}" for i in range(len(cols), width)])[:width]
    df = pd.DataFrame(data_rows, columns=use_cols)
    df = df.dropna(how="all")
    if df.empty:
        return df
    df["Sondage"] = df["Sondage"].ffill()
    for c in ["Easting_rep", "Northing_rep", "Elevation_rep", "Easting_man", "Northing_man", "Elevation_man"]:
        if c in df.columns:
            df[c] = df.groupby("Sondage")[c].transform(lambda s: s.ffill().bfill())
    df["Easting"] = df["Easting_rep"].combine_first(df["Easting_man"])
    df["Northing"] = df["Northing_rep"].combine_first(df["Northing_man"])
    df["Elevation"] = df["Elevation_rep"].combine_first(df["Elevation_man"])
    df["Azimut"] = pd.to_numeric(df.get("Azimut"), errors="coerce")
    df["Pendage"] = pd.to_numeric(df.get("Pendage"), errors="coerce")
    df["From_m"] = pd.to_numeric(df.get("From_m"), errors="coerce")
    df["To_m"] = pd.to_numeric(df.get("To_m"), errors="coerce")

    def dip_direction(az):
        if pd.isna(az):
            return None
        sectors = ["N", "NNE", "NE", "ENE", "E", "ESE", "SE", "SSE", "S", "SSW", "SW", "WSW", "W", "WNW", "NW", "NNW"]
        idx = int((az % 360) / 22.5 + 0.5) % 16
        return sectors[idx]

    df["Sens_Pendage"] = df["Azimut"].apply(dip_direction)
    return df


def generate_demo_data():
    """Génère un jeu de données synthétique (démonstration) pour visualiser le dashboard
    avant l'import de données réelles complètes."""
    rng = np.random.default_rng(42)
    drill_defs = {
        "RC": [("RC-001", 450000, 850000, 320), ("RC-002", 450050, 850010, 322), ("RC-003", 450100, 850020, 325)],
        "AC": [("AC-001", 450010, 850060, 321), ("AC-002", 450060, 850070, 323)],
        "DD": [("DD-001", 450005, 850005, 320), ("DD-002", 450090, 850015, 326)],
    }
    lithos = ["Saprolite argileux", "Volcanique mafique altéré", "Volcanoclastite", "Granitoïde",
              "Schiste graphiteux", "BIF (Formation de fer)", "Diorite", "Quartzite"]
    strat = [("SUP", "Superficiel"), ("LAT", "Latérite"), ("SAP", "Saprolite"), ("SAPROCK", "Saprock"),
             ("VOLC1", "Volcanique Inf"), ("FRESH", "Roche fraîche/socle")]
    out = {}
    for dtype, holes in drill_defs.items():
        rows = []
        for hole, e, n, elev in holes:
            depth = rng.integers(60, 140)
            cur = 0.0
            while cur < depth:
                step = rng.integers(3, 12)
                to = min(cur + step, depth)
                s_code, s_name = strat[min(len(strat) - 1, int(cur / depth * len(strat)))]
                litho = rng.choice(lithos)
                has_min = rng.random() < 0.25
                has_alt = rng.random() < 0.35
                rows.append({
                    "Easting": e, "Northing": n, "Elevation": elev, "Sondage": hole,
                    "From": cur, "To": to, "Lithologie": litho, "Code_Litho": litho[:4].upper(),
                    "Code_Strat": s_code, "Formation": s_name, "Texture": rng.choice(["Massive", "Foliée", "Bréchique"]),
                    "Weathering_Class": s_code, "Has_Mineralisation": has_min, "Has_Alteration": has_alt,
                    "Has_Veine": rng.random() < 0.15, "Drill_Type": dtype,
                })
                cur = to
        out[dtype] = pd.DataFrame(rows)

    # données structurales synthétiques (mesures réparties dans les trous DD/RC)
    struct_rows = []
    all_holes = [h for d in drill_defs.values() for h in d]
    for hole, e, n, elev in all_holes:
        depth = out["RC"]["To"].max() if not out["RC"].empty else 100
        n_meas = rng.integers(4, 10)
        for _ in range(n_meas):
            depth_pt = rng.uniform(5, 110)
            az = rng.normal(135, 35) % 360
            dip = abs(rng.normal(55, 15))
            dip = min(max(dip, 5), 89)
            struct_rows.append({
                "Sondage": hole, "Easting": e, "Northing": n, "Elevation": elev,
                "From_m": depth_pt, "To_m": depth_pt + 0.1, "Azimut": az, "Pendage": dip,
                "Sens_Pendage": None, "OCTypes": rng.choice(["Foliation", "Fracture", "Veine", "Faille", "Contact"]),
                "Qualite": rng.choice(["Bonne", "Moyenne", "Faible"]),
                "GSI_auto": min(max(rng.normal(45, 18), 5), 95),
                "RQD_auto": min(max(rng.normal(55, 20), 0), 100),
            })
    sdf = pd.DataFrame(struct_rows)

    def dip_direction(az):
        sectors = ["N", "NNE", "NE", "ENE", "E", "ESE", "SE", "SSE", "S", "SSW", "SW", "WSW", "W", "WNW", "NW", "NNW"]
        idx = int((az % 360) / 22.5 + 0.5) % 16
        return sectors[idx]
    sdf["Sens_Pendage"] = sdf["Azimut"].apply(dip_direction)
    out["STRUCT"] = sdf

    # géochimie sols / auger synthétique avec teneurs Au
    auger_rows = []
    for i in range(1, 9):
        e, n = 449980 + i * 25, 850040 + rng.integers(-10, 10)
        depth = 0.0
        for _ in range(rng.integers(2, 5)):
            to = depth + rng.uniform(0.5, 1.5)
            au = max(0, rng.lognormal(mean=2.5, sigma=1.2))  # ppb
            auger_rows.append({
                "Easting": e, "Northing": n, "Elevation": 318 + rng.integers(-5, 5), "Sondage": f"AG-{i:03d}",
                "From": depth, "To": to, "Lithologie": rng.choice(["Sol organique", "Latérite", "Saprolite argileux"]),
                "Weathering_Class": rng.choice(["SUP", "LAT", "SAP"]), "Au_ppb": au, "Drill_Type": "AUGER",
            })
            depth = to
    out["AUGER"] = pd.DataFrame(auger_rows)
    return out


def hole_trajectory(easting, northing, elevation, azimut, pendage, depth, az_drift=0.0, dip_drift=0.0, n_points=40):
    """Calcule la trajectoire 3D d'un trou de forage (modèle simplifié de désurvey).
    azimut : direction de visée en degrés (0=N, 90=E). pendage : inclinaison sous l'horizontale en degrés
    (90 = vertical). az_drift / dip_drift : dérive cumulée en degrés sur toute la longueur du trou,
    simulant une déviation progressive (ex: torsion du train de tige, désurvey gyroscopique).
    Retourne un DataFrame avec colonnes Depth, Easting, Northing, Elevation, Azimut_loc, Pendage_loc."""
    if depth is None or depth <= 0 or pd.isna(depth):
        depth = 50.0
    if azimut is None or pd.isna(azimut):
        azimut = 0.0
    if pendage is None or pd.isna(pendage):
        pendage = 90.0
    depths = np.linspace(0, depth, n_points)
    pts = []
    for d in depths:
        frac = d / depth if depth else 0
        az_loc = azimut + az_drift * frac
        dip_loc = pendage - dip_drift * frac  # le pendage tend à s'aplatir (effet gravité) si dip_drift > 0
        dip_loc = min(max(dip_loc, 1), 90)
        az_rad = np.radians(az_loc)
        dip_rad = np.radians(dip_loc)
        horiz = d * np.cos(dip_rad)
        dz = -d * np.sin(dip_rad)
        dx = horiz * np.sin(az_rad)
        dy = horiz * np.cos(az_rad)
        pts.append({
            "Depth": d, "Easting": easting + dx, "Northing": northing + dy,
            "Elevation": elevation + dz, "Azimut_loc": az_loc, "Pendage_loc": dip_loc,
        })
    return pd.DataFrame(pts)


def length_weighted_grade(df, value_col="Au_ppb", from_col="From", to_col="To", group_col="Sondage"):
    """Calcule la teneur moyenne pondérée par la longueur d'intervalle, par trou."""
    if df is None or df.empty or value_col not in df.columns:
        return pd.DataFrame()
    d = df.copy()
    d[value_col] = pd.to_numeric(d[value_col], errors="coerce")
    d[from_col] = pd.to_numeric(d[from_col], errors="coerce")
    d[to_col] = pd.to_numeric(d[to_col], errors="coerce")
    d = d.dropna(subset=[value_col, from_col, to_col])
    if d.empty:
        return pd.DataFrame()
    d["Longueur"] = d[to_col] - d[from_col]
    d["Longueur"] = d["Longueur"].clip(lower=0.01)
    d["Pondere"] = d[value_col] * d["Longueur"]
    g = d.groupby(group_col).apply(
        lambda x: pd.Series({
            "Longueur_totale_m": x["Longueur"].sum(),
            "Teneur_moy_ppb": x["Pondere"].sum() / x["Longueur"].sum(),
            "Teneur_max_ppb": x[value_col].max(),
            "N_intervalles": len(x),
        }), include_groups=False
    ).reset_index()
    g["Teneur_moy_g_t"] = g["Teneur_moy_ppb"] / 1000.0
    g["Teneur_max_g_t"] = g["Teneur_max_ppb"] / 1000.0
    return g


def audit_dataframe(df, label, hole_col="Sondage", from_col="From", to_col="To"):
    """Audite un DataFrame de log et retourne une liste d'anomalies détectées, avec
    suggestion de correction. Chaque anomalie est un dict {severite, source, trou, message, fix}."""
    issues = []
    if df is None or df.empty:
        return issues
    fcol = from_col if from_col in df.columns else None
    tcol = to_col if to_col in df.columns else None

    # 1. coordonnées manquantes
    if "Easting" in df.columns and hole_col in df.columns:
        collars = df.groupby(hole_col)["Easting"].apply(lambda s: s.notna().any())
        for h, ok in collars.items():
            if not ok:
                issues.append({"severite": "⚠️ Erreur", "source": label, "trou": h,
                                "message": "Coordonnées (Easting/Northing) manquantes pour ce trou.",
                                "fix": "Saisir les coordonnées collar dans le fichier source."})

    if fcol and tcol:
        d = df.copy()
        d[fcol] = pd.to_numeric(d[fcol], errors="coerce")
        d[tcol] = pd.to_numeric(d[tcol], errors="coerce")

        # 2. From > To
        bad = d[d[fcol] > d[tcol]]
        for _, r in bad.iterrows():
            issues.append({"severite": "🔴 Critique", "source": label, "trou": r.get(hole_col, "?"),
                            "message": f"Intervalle invalide : From ({r[fcol]}) > To ({r[tcol]}).",
                            "fix": "Inverser ou corriger From/To."})

        # 3. profondeurs négatives
        neg = d[(d[fcol] < 0) | (d[tcol] < 0)]
        for _, r in neg.iterrows():
            issues.append({"severite": "🔴 Critique", "source": label, "trou": r.get(hole_col, "?"),
                            "message": "Profondeur négative détectée.",
                            "fix": "Vérifier la saisie From/To (valeur négative impossible)."})

        # 4. chevauchements et trous (gaps) par sondage
        if hole_col in d.columns:
            for h, g in d.dropna(subset=[fcol, tcol]).groupby(hole_col):
                g = g.sort_values(fcol)
                prev_to = None
                for _, r in g.iterrows():
                    if prev_to is not None:
                        if r[fcol] < prev_to - 0.01:
                            issues.append({"severite": "⚠️ Erreur", "source": label, "trou": h,
                                            "message": f"Chevauchement d'intervalles autour de {r[fcol]}-{prev_to} m.",
                                            "fix": "Trier/recouper les intervalles pour éliminer le chevauchement."})
                        elif r[fcol] > prev_to + 0.01:
                            issues.append({"severite": "ℹ️ Info", "source": label, "trou": h,
                                            "message": f"Lacune (gap) non loggée entre {prev_to} et {r[fcol]} m.",
                                            "fix": "Vérifier si l'intervalle manquant doit être renseigné."})
                    prev_to = r[tcol]

    # 5. lithologie manquante
    if "Lithologie" in df.columns:
        miss = df["Lithologie"].isna().sum()
        if miss > 0:
            issues.append({"severite": "ℹ️ Info", "source": label, "trou": "Plusieurs",
                            "message": f"{miss} intervalle(s) sans lithologie renseignée.",
                            "fix": "Compléter le champ Lithologie pour ces intervalles."})

    return issues


def auto_fix_dataframe(df, hole_col="Sondage", from_col="From", to_col="To"):
    """Corrections automatiques sûres : tri des intervalles par trou/From, suppression des
    doublons stricts, et échange From/To si From > To (corrections strictement mécaniques —
    aucune valeur n'est inventée)."""
    if df is None or df.empty:
        return df
    d = df.copy()
    if from_col in d.columns and to_col in d.columns:
        d[from_col] = pd.to_numeric(d[from_col], errors="coerce")
        d[to_col] = pd.to_numeric(d[to_col], errors="coerce")
        swap_mask = d[from_col] > d[to_col]
        d.loc[swap_mask, [from_col, to_col]] = d.loc[swap_mask, [to_col, from_col]].values
        if hole_col in d.columns:
            d = d.sort_values([hole_col, from_col], na_position="last").reset_index(drop=True)
    d = d.drop_duplicates()
    return d


def generate_geological_report(data, litho_colors, prospect="Prospect", permis=""):
    """Compile un rapport géologique structuré (texte/Markdown) à partir de toutes les données
    chargées dans le dashboard : contexte, lithologie, minéralisation, altération, structures,
    interprétation et recommandations."""
    all_df = pd.concat([data.get(k, pd.DataFrame()) for k in ["RC", "AC", "DD"] if not data.get(k, pd.DataFrame()).empty],
                        ignore_index=True) if any(not data.get(k, pd.DataFrame()).empty for k in ["RC", "AC", "DD"]) else pd.DataFrame()
    sdf = data.get("STRUCT", pd.DataFrame())
    auger = data.get("AUGER", pd.DataFrame())

    sections = {}
    sections["Résumé exécutif"] = (
        f"Ce rapport synthétise les données d'exploration disponibles pour le prospect **{prospect}**"
        + (f" (permis {permis})" if permis else "") + f", générées automatiquement à partir de "
        f"{0 if all_df.empty else all_df['Sondage'].nunique()} trou(s) de forage (RC/AC/DD) et "
        f"{0 if auger.empty else auger['Sondage'].nunique()} point(s) Auger/géochimie sols. "
        f"Il constitue une base d'interprétation rapide ; il ne remplace pas un rapport technique "
        f"complet rédigé et validé par une Personne Qualifiée."
    )

    if not all_df.empty:
        collars = collar_table(all_df)
        litho_counts = all_df["Lithologie"].dropna().value_counts() if "Lithologie" in all_df.columns else pd.Series(dtype=int)
        n_mineral_holes = int(all_df.groupby("Sondage")["Has_Mineralisation"].any().sum()) if "Has_Mineralisation" in all_df.columns else 0
        total_m = collars["Profondeur_totale"].sum() if "Profondeur_totale" in collars.columns else 0
        txt = (f"Le programme de forage couvre **{len(collars)} trou(s)** pour un total de "
               f"**{total_m:.0f} m** forés. ")
        if not litho_counts.empty:
            top3 = ", ".join([f"{l} ({c} interv.)" for l, c in litho_counts.head(3).items()])
            txt += f"Les lithologies les plus fréquemment recoupées sont : {top3}. "
        txt += (f"**{n_mineral_holes} trou(s)** sur {len(collars)} (soit "
                f"{n_mineral_holes / max(len(collars), 1) * 100:.0f}%) présentent des indices de "
                f"minéralisation sulfurée d'après les logs disponibles.")
        sections["Contexte géologique & lithologie"] = txt
    else:
        sections["Contexte géologique & lithologie"] = "Aucune donnée RC/AC/DD chargée — section non générée."

    if not all_df.empty and "Has_Alteration" in all_df.columns:
        n_alt = int(all_df["Has_Alteration"].sum())
        pct_alt = n_alt / len(all_df) * 100 if len(all_df) else 0
        sections["Altération & minéralisation"] = (
            f"Sur l'ensemble des intervalles logués, **{n_alt} ({pct_alt:.0f}%)** présentent une "
            f"altération hydrothermale renseignée. La coexistence fréquente d'altération et de "
            f"minéralisation sulfurée (à vérifier trou par trou dans l'onglet Logs automatisés) est "
            f"typique d'un système hydrothermal actif — orogénique ou épithermal selon le contexte "
            f"régional. Il est recommandé de qualifier précisément les assemblages d'altération "
            f"(séricite-chlorite, argilique, propylitique, potassique) pour positionner le système "
            f"dans un modèle génétique reconnu."
        )

    if not sdf.empty and "Azimut" in sdf.columns and sdf["Azimut"].dropna().any():
        az_mean = sdf["Azimut"].mean()
        dip_mean = sdf["Pendage"].mean()
        sections["Structures"] = (
            f"**{len(sdf)} mesure(s) structurale(s)** sont disponibles, avec un azimut moyen de "
            f"**{az_mean:.0f}°** et un pendage moyen de **{dip_mean:.0f}°**. Cette orientation "
            f"dominante doit être comparée à la fabrique régionale connue pour distinguer la "
            f"foliation/schistosité principale d'éventuelles structures minéralisantes tardives "
            f"(failles, veines, zones de cisaillement) qui contrôlent souvent la distribution de l'or "
            f"en contexte orogénique."
        )

    if not auger.empty and "Au_ppb" in auger.columns and auger["Au_ppb"].notna().any():
        au_max = auger["Au_ppb"].max()
        au_mean = auger["Au_ppb"].mean()
        seuil = auger["Au_ppb"].quantile(0.75)
        n_anom = int((auger["Au_ppb"] >= seuil).sum())
        sections["Géochimie sols / Auger"] = (
            f"La géochimie sols indique une teneur Au moyenne de **{au_mean:.1f} ppb** (max "
            f"**{au_max:.1f} ppb**). **{n_anom} échantillon(s)** dépassent le seuil anomal "
            f"(75e percentile, {seuil:.1f} ppb), définissant des cibles prioritaires pour un "
            f"programme de forage RC/AC de vérification en profondeur."
        )

    recommendations = []
    if not all_df.empty:
        n_holes = all_df["Sondage"].nunique()
        if n_holes < 10:
            recommendations.append("Le nombre de trous est encore limité (<10) : un programme d'infill "
                                    "est recommandé pour densifier la grille avant toute estimation de "
                                    "ressources classée.")
    if not sdf.empty and sdf["Azimut"].dropna().empty:
        recommendations.append("Compléter les mesures structurales (azimut/pendage) sur l'ensemble des "
                                "trous DD pour fiabiliser l'interprétation du contrôle structural.")
    if auger.empty:
        recommendations.append("Aucune donnée géochimie sols/Auger chargée : un levé géochimique "
                                "systématique permettrait de mieux cibler les zones d'extension latérale.")
    recommendations.append("Mettre en place un programme QAQC rigoureux (≥10% d'échantillons de "
                            "contrôle) avant toute estimation de ressources publiable.")
    recommendations.append("Valider les limites d'altération météorique (latérite/saprolite/saprock/"
                            "socle) par photo-interprétation ou levé de terrain complémentaire pour "
                            "fiabiliser les corrélations de section.")
    sections["Recommandations"] = "\n".join([f"- {r}" for r in recommendations])

    return sections


def report_to_markdown(sections, prospect, permis):
    md = f"# Rapport géologique — {prospect}" + (f" (permis {permis})" if permis else "") + "\n\n"
    md += f"*Généré automatiquement par ESPACE VIRTUELLE MINIÈRE DE SMC*\n\n---\n\n"
    for title, text in sections.items():
        md += f"## {title}\n\n{text}\n\n"
    return md


def collar_table(df, depth_col_candidates=("To", "To_m")):
    """Construit une table des collars (1 ligne par trou) : easting, northing, elevation,
    profondeur totale, type de forage."""
    if df is None or df.empty:
        return pd.DataFrame()
    depth_col = None
    for c in depth_col_candidates:
        if c in df.columns:
            depth_col = c
            break
    agg = {"Easting": "first", "Northing": "first", "Elevation": "first"}
    if "Drill_Type" in df.columns:
        agg["Drill_Type"] = "first"
    g = df.groupby("Sondage").agg(agg)
    if depth_col:
        g["Profondeur_totale"] = df.groupby("Sondage")[depth_col].max()
    g = g.reset_index()
    return g
    

